import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import filedialog, Tk

def to_float(s):
    try:
        s = str(s).strip()
        if s == "":
            return 0.0
        return float(s.replace(".", "").replace(",", "."))
    except:
        return 0.0

def split_normalized(line):
    parts = line.split("|")
    if len(parts) and parts[0] == "":
        parts = parts[1:]
    return parts

def analisar_sped(caminho):
    with open(caminho, "r", encoding="latin-1") as f:
        linhas = [ln.rstrip("\n") for ln in f]

    resultados = []
    i = 0
    while i < len(linhas):
        line = linhas[i]
        parts = split_normalized(line)

        if len(parts) > 0 and parts[0] == "C100":
            numero = parts[7] if len(parts) > 7 else ""
            valor_total = to_float(parts[11]) if len(parts) > 11 else 0.0

            soma_produtos = 0.0
            i += 1
            while i < len(linhas) and not linhas[i].startswith("|C100|"):
                parts2 = split_normalized(linhas[i])
                if len(parts2) > 0 and parts2[0] == "C190":
                    soma_produtos += to_float(parts2[4]) if len(parts2) > 4 else 0.0
                i += 1

            resultado = "Correto" if abs(soma_produtos - valor_total) < 0.01 else "Erro"
            resultados.append([numero, soma_produtos, valor_total, resultado])
        else:
            i += 1

    salvar_excel(resultados, caminho)

def salvar_excel(resultados, caminho_entrada):
    pasta_saida = os.path.dirname(caminho_entrada)
    nome_saida = os.path.join(pasta_saida, "Conferencia_SPED.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conferência SPED"

    # Cabeçalho
    headers = ["Nota", "Soma C190", "Valor Total Nota", "Resultado"]
    ws.append(headers)

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Inserir dados
    for row in resultados:
        ws.append(row)

    # Formatação condicional
    for i, row in enumerate(resultados, start=2):
        cell_resultado = ws.cell(row=i, column=4)
        texto = str(row[3]).strip().lower()
        if texto == "correto":
            cell_resultado.font = Font(color="008000", bold=True)  # verde
        elif texto == "erro":
            cell_resultado.font = Font(color="FF0000", bold=True)  # vermelho

        # Converter Nota para inteiro
        try:
            ws.cell(row=i, column=1).value = int(float(row[0]))
        except:
            pass

    # Ajustar largura automática
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4

    # Adicionar filtros
    ws.auto_filter.ref = ws.dimensions

    wb.save(nome_saida)
    print(f"\n✅ Planilha salva em: {nome_saida}\n")

def selecionar_arquivo():
    root = Tk()
    root.withdraw()  # esconde janela principal
    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo SPED",
        filetypes=[("Arquivos TXT", "*.txt")]
    )
    root.destroy()

    if caminho:
        print(f"\n📁 Arquivo selecionado: {caminho}\n")
        analisar_sped(caminho)
    else:
        print("⚠ Nenhum arquivo selecionado.")

if __name__ == "__main__":
    selecionar_arquivo()
